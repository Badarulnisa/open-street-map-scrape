"""
Excel export. The single place responsible for turning Business objects
into the final .xlsx deliverable(s). Collectors and the merger never touch
pandas/openpyxl directly — they just produce/combine Business objects
and hand them here.
"""
from pathlib import Path

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import config
from models.business import Business
from utils.logger import get_logger

logger = get_logger(__name__)

# --- Raw export (full dataset, no enrichment) ------------------------------
EXPORT_COLUMNS: list[str] = [
    "name", "city", "area", "category", "website", "phone",
    "linkedin_url", "latitude", "longitude", "source",
    "verification_status", "confidence_score",
]

# --- Enriched export (demo subset — includes all enrichment fields) ------
ENRICHED_EXPORT_COLUMNS: list[str] = [
    "name", "city", "area", "category", "address", "latitude", "longitude",
    "phone", "source",
    "website_verified", "website_confidence",
    "linkedin_url", "linkedin_confidence",
    "google_maps_url",
    "facebook_url", "instagram_url", "twitter_url", "youtube_url", "crunchbase_url",
    "social_confidence",
    "email", "company_description", "industry", "headquarters",
    "company_size", "founded", "careers_url",
    "registration_status", "trade_license_authority", "registration_confidence",
    "verification_status", "confidence_score",
    "search_status", "search_notes",
]

# Confidence-score columns that get green/yellow/red conditional formatting
CONFIDENCE_COLUMNS: list[str] = [
    "website_confidence", "linkedin_confidence", "social_confidence",
    "registration_confidence",
]


def _apply_base_formatting(worksheet: Worksheet, df: pd.DataFrame) -> None:
    """Bold header, frozen header row, autofilter, auto-width columns.
    Shared by both raw and enriched exports so formatting never drifts
    out of sync between the two files."""
    # Bold header row
    for cell in worksheet[1]:
        cell.font = cell.font.copy(bold=True)

    # Freeze header row
    worksheet.freeze_panes = "A2"

    # Autofilter across the full data range
    worksheet.auto_filter.ref = worksheet.dimensions

    # Auto-adjust column widths based on content length.
    # Uses .apply() with an explicit pd.notna() check rather than
    # .astype(str).map(len), because NaN float cells (common in columns
    # with missing enrichment data) crash .map(len) with a TypeError.
    for i, col in enumerate(df.columns, start=1):
        if df.empty:
            content_max = 0
        else:
            content_max = df[col].apply(
                lambda x: len(str(x)) if pd.notna(x) else 0
            ).max()
        max_len = max(content_max, len(col))
        worksheet.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 60)


def save_to_excel(businesses: list[Business], output_path: Path) -> Path:
    """Write a list of Business objects to a single Excel file at output_path.
    Used for the raw, full-dataset export (no enrichment)."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if not businesses:
        logger.warning(f"No businesses to write to {output_path}.")

    rows = [b.to_dict() for b in businesses]
    df = pd.DataFrame(rows).reindex(columns=EXPORT_COLUMNS)

    with pd.ExcelWriter(output_path, engine=config.EXCEL_ENGINE) as writer:
        df.to_excel(writer, index=False, sheet_name="Dubai")
        worksheet = writer.sheets["Dubai"]
        _apply_base_formatting(worksheet, df)

    logger.info(f"Saved: {output_path} ({len(businesses)} rows)")
    return output_path


def save_enriched_excel(businesses: list[Business], output_path: Path = None) -> Path:
    """Write enriched Business objects to a formatted Excel file: bold
    headers, frozen header row, autofilter, auto-width columns, and
    green/yellow/red conditional formatting on confidence scores."""
    output_path = output_path or config.ENRICHMENT_OUTPUT_PATH
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if not businesses:
        logger.warning(f"No businesses to write to {output_path}.")

    rows = [b.to_dict() for b in businesses]
    df = pd.DataFrame(rows).reindex(columns=ENRICHED_EXPORT_COLUMNS)

    with pd.ExcelWriter(output_path, engine=config.EXCEL_ENGINE) as writer:
        df.to_excel(writer, index=False, sheet_name="Dubai_Enriched")
        worksheet = writer.sheets["Dubai_Enriched"]
        _apply_base_formatting(worksheet, df)

        # Conditional formatting on confidence columns: green >=70, yellow 30-69, red <30
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        n_rows = len(df) + 1  # +1 accounts for the header row
        for col_name in CONFIDENCE_COLUMNS:
            if col_name not in df.columns:
                continue
            col_idx = df.columns.get_loc(col_name) + 1
            col_letter = get_column_letter(col_idx)
            cell_range = f"{col_letter}2:{col_letter}{n_rows}"

            worksheet.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="greaterThanOrEqual", formula=["70"], fill=green_fill),
            )
            worksheet.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="between", formula=["30", "69"], fill=yellow_fill),
            )
            worksheet.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="lessThan", formula=["30"], fill=red_fill),
            )

    logger.info(f"Saved enriched export: {output_path} ({len(businesses)} rows)")
    return output_path